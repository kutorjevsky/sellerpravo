"""
Конвертер старого Excel-файла в CSV формата нового реестра.

Берёт лист с завершёнными поручениями (по умолчанию «ИСПОЛНЕННОЕ») и приводит
10 старых колонок к колонкам нового реестра. Результат — CSV, который удобно
импортировать в Google-таблицу: Файл → Импорт → Загрузить → вставить в новый
лист (например, 99_Архив).

Старые рабочие листы «на каждый день» не переносятся: это снимки-дубли одних и тех
же задач, переходящих изо дня в день. Реальная история — именно лист ИСПОЛНЕННОЕ.

Запуск:
    python tools/import_old.py путь/к/файлу.xlsx [ИмяЛиста] [out.csv]
"""
import csv
import sys
import warnings
from datetime import datetime

import openpyxl

from bot import config as cfg

warnings.filterwarnings("ignore")


def norm_poa(text: str) -> str:
    t = (text or "").lower()
    if "нет" in t:
        return "Нет"
    if "дов" in t:
        return "Да"
    return ""


def fmt_date(val) -> str:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val or "").strip()


def convert(path: str, sheet_name: str = "ИСПОЛНЕННОЕ", out: str = "archive.csv") -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Лист «{sheet_name}» не найден. Есть: {wb.sheetnames[:10]}…")
    ws = wb[sheet_name]

    rows_out = []
    seq = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # заголовок
        cells = list(row) + [None] * 10
        date, initiator, urgency, organ, client, task, courier_comment, \
            courier, marks, poa_status = cells[:10]

        # пропускаем пустые и строки-заметки (есть текст только в 1-й колонке)
        if not any([client, task, courier, courier_comment]):
            continue

        d = fmt_date(date)
        prefix = d if len(d) == 10 else "архив"
        seq[prefix] = seq.get(prefix, 0) + 1
        task_id = f"{prefix}-{seq[prefix]:03d}"

        comments = []
        if marks:
            comments.append(f"[отметки] {marks}")
        status = cfg.ST_ACCEPTED if courier_comment else cfg.ST_CANCELLED

        rec = {
            cfg.H_ID: task_id,
            cfg.H_CREATED: d,
            cfg.H_STATUS: status,
            cfg.H_INITIATOR: str(initiator or "").strip(),
            cfg.H_CLIENT: str(client or "").strip(),
            cfg.H_ORGAN: str(organ or "").strip(),
            cfg.H_ACTION: str(task or "").strip()[:120],
            cfg.H_FULL: str(task or "").strip(),
            cfg.H_LEGAL_DEADLINE: str(urgency or "").strip(),
            cfg.H_POA: norm_poa(poa_status),
            cfg.H_EXECUTOR: str(courier or "").strip(),
            cfg.H_RESULT: str(courier_comment or "").strip(),
            cfg.H_COMMENTS: "\n".join(comments),
        }
        rows_out.append([rec.get(h, "") for h in cfg.REGISTRY_HEADERS])

    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(cfg.REGISTRY_HEADERS)
        w.writerows(rows_out)

    print(f"Готово: {len(rows_out)} записей → {out}")
    return len(rows_out)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python tools/import_old.py файл.xlsx [Лист] [out.csv]")
        raise SystemExit(1)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else "ИСПОЛНЕННОЕ"
    out = sys.argv[3] if len(sys.argv) > 3 else "archive.csv"
    convert(path, sheet, out)
